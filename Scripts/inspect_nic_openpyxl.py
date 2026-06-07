import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = "Scripts/jmp_nicaragua_household.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

print("Sheet names:", wb.sheetnames)
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"Sheet: {sheet_name}, max_row={sheet.max_row}, max_column={sheet.max_column}")
