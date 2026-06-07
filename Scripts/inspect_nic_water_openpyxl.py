import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = "Scripts/jmp_nicaragua_household.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Water Data']

print(f"Sheet dimensions: max_row={sheet.max_row}, max_column={sheet.max_column}")

# Print first 10 rows and 10 columns
for r in range(1, 15):
    row_vals = []
    for c in range(1, 15):
        row_vals.append(sheet.cell(row=r, column=c).value)
    # Check if the row has any values
    if any(x is not None for x in row_vals):
        print(f"Row {r}: {row_vals}")
